from selenium import webdriver
from selenium.webdriver.common.by import By
import datetime
import time
import os
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import sys
from selenium.common.exceptions import TimeoutException

from multiprocessing import Queue, Process
import openpyxl
import xlsxwriter
from selenium.common.exceptions import NoSuchElementException as NoElement
from selenium.common.exceptions import NoSuchWindowException as NoWindow
from selenium.common.exceptions import WebDriverException as Chrome_not_reachable
# url = 'https://e-dostavka.by/catalog/item_766362.html'
path = 'D:\python\e-dostavka.by\\'


class Parser:
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.excel_filename = 'products.xlsx'
        self.writed_urls = []
        self.excel_counter = 2

    def start(self, url):
        self.driver.get(url)
        if len(self.driver.find_elements_by_class_name('content'))==1:
            self.is_remove = True
        self.name = 'None'
        self.categories_trees_list = []
        self.photo = 'None'
        self.price = 'None'
        self.old_price = ''
        self.article = 'None'
        self.barcode = 'None'
        self.country = 'None'
        self.trademark = 'None'
        self.weight = 'None'
        self.property = ''
        self.about = 'None'
        self.maker = 'None'
        # self.excel_counter = 2
        self.get_name()
        self.get_categories_trees()
        self.get_photo()
        self.get_price()
        self.get_description()
        self.get_property()
        self.get_about()
        self.get_characteristics()
        self.write_to_excel()


    def get_name(self):
        self.name = self.driver.find_element_by_tag_name('h1').text

    def get_categories_trees(self):
        self.tree_values = self.driver.find_element_by_class_name('breadcrumbs').find_elements_by_tag_name('a')
        for value in self.tree_values:
            self.categories_trees_list.append(value.text)
        self.categories_trees_list.pop(0)
        self.categories_trees_list.pop(0)

    def get_photo(self):
        self.photo = self.driver.find_element_by_class_name('increaseImage').get_attribute('href')

    def get_price(self):
        try:
            # print('цена self.price до: ', self.price)
            self.price = self.driver.find_element_by_class_name('product_card__inner').find_element_by_class_name('price').text.split('\n')[0]
            # print('цена self.price после: ', self.price)
            # print('цена self.old_price до: ', self.old_price)

            self.old_price = self.driver.find_element_by_class_name('product_card__inner').find_element_by_class_name('old_price').text
            # print('цена self.old_price после: ', self.old_price)
        except NoElement:
            self.old_price = ''


    def get_description(self): # Article, barcode, country, trademark, weight
        description_table = self.driver.find_element_by_class_name('description').find_elements_by_tag_name('li')
        description_dict = {}
        for li in description_table:
            description_dict[li.find_element_by_tag_name('strong').text] = li.find_element_by_tag_name('span').text
        try:
            self.article = description_dict['Артикул:']
        except KeyError:
            pass
        try:
            self.barcode = description_dict['Штрих-код:']
        except KeyError:
            pass
        try:
            self.country = description_dict['Страна производства:']
        except KeyError:
            pass
        try:
            self.trademark = description_dict['Торговая марка:']
        except KeyError:
            pass
        try:
            self.weight = (description_dict['Масса:']).split(' ')[0]
            self.unit = (description_dict['Масса:']).split(' ')[-1]
        except KeyError:
            pass

    def get_property(self):
        try:
            property_table = self.driver.find_element_by_class_name('property_group_9').find_elements_by_tag_name('tr')
            # property_dict = {}
            for tr in property_table:
                # property_dict[tr.find_element_by_class_name('name').text] = tr.find_element_by_class_name('value').text
                self.property+=tr.text
                self.property+='\n'
        except NoElement:
            pass

    def get_about(self):
        try:
            self.about = self.driver.find_element_by_class_name('property_3220').text
            print(self.about)
        except NoElement:
            try:
                self.about = self.driver.find_element_by_class_name('property_3221').text
                print(self.about)
            except NoElement:
                pass


    def get_characteristics(self):
        time.sleep(0.5)
        self.driver.find_element_by_xpath('//li[@class="ui-state-default ui-corner-top"]').click()
        self.characteristics = str()
        characteristics_table = self.driver.find_element_by_id('tab-1').find_elements_by_tag_name('tr')
        for tr in characteristics_table:
            self.characteristics += tr.find_element_by_class_name('name').text
            self.characteristics += '\n'
            self.characteristics +=tr.find_element_by_class_name('value').text
            self.characteristics += '\n'
            if tr.find_element_by_class_name('name').text == 'Производитель':
                self.maker = tr.find_element_by_class_name('value').text.split(',')[0]
        print(self.characteristics)

    def try_make_new_excelbook(self):
        if self.excel_filename not in os.listdir(path):
            self.workbook = xlsxwriter.Workbook(path + self.excel_filename)
            self.worksheet = self.workbook.add_worksheet('goods')
            self.workbook.close()
            self._prepare_excel_sheet()
            return []
        else:
            return self._chk_urls()

    def _chk_urls(self):
        self.workbook = openpyxl.load_workbook(path + self.excel_filename)
        self.worksheet = self.workbook['goods']
        self.excel_counter = 1
        cell = self.worksheet.cell(row=self.excel_counter, column=21).value
        while cell:
            self.excel_counter += 1
            cell = self.worksheet.cell(row=self.excel_counter, column=21).value
            if cell:
                self.writed_urls.append(cell.replace('\n', ''))

        return self.writed_urls

    def _prepare_excel_sheet(self):
        self.workbook = openpyxl.load_workbook(path + self.excel_filename)
        self.worksheet = self.workbook['goods']
        self.worksheet['A1'] = 'Product name'
        self.worksheet['B1'] = 'Category 1'
        self.worksheet['C1'] = 'Category 2'
        self.worksheet['D1'] = 'Category 3'
        self.worksheet['E1'] = 'Category 4'
        self.worksheet['F1'] = 'Category 5'
        self.worksheet['G1'] = 'Category 6'
        self.worksheet['H1'] = 'Article'
        self.worksheet['I1'] = 'barcode'
        self.worksheet['J1'] = 'country'
        self.worksheet['K1'] = 'trademark'
        self.worksheet['L1'] = 'Produced by:'
        self.worksheet['M1'] = 'weight'
        self.worksheet['N1'] = 'unit'
        self.worksheet['O1'] = 'price'
        self.worksheet['P1'] = 'old price'
        self.worksheet['Q1'] = 'characteristics'
        self.worksheet['R1'] = 'property'
        self.worksheet['S1'] = 'about'
        self.worksheet['T1'] = 'Photo'
        self.worksheet['U1'] = 'Link'
        self.worksheet['V1'] = 'Date of renew'


        self.workbook.save(path + self.excel_filename)

    def write_to_excel(self):
        # self.workbook = openpyxl.load_workbook(path + self.excel_filename)
        # self.worksheet = self.workbook['goods']

        cell = self.worksheet.cell(row=self.excel_counter, column=1).value
        while cell:
            self.excel_counter += 1
            cell = self.worksheet.cell(row=self.excel_counter, column=21).value
        self.worksheet['A' + str(self.excel_counter)] = self.name
        try:
            self.worksheet['B' + str(self.excel_counter)] = self.categories_trees_list[0]
            self.worksheet['C' + str(self.excel_counter)] = self.categories_trees_list[1]
            self.worksheet['D' + str(self.excel_counter)] = self.categories_trees_list[2]
            self.worksheet['E' + str(self.excel_counter)] = self.categories_trees_list[3]
            self.worksheet['F' + str(self.excel_counter)] = self.categories_trees_list[4]
            self.worksheet['G' + str(self.excel_counter)] = self.categories_trees_list[5]
        except IndexError:
            pass

        self.worksheet['H' + str(self.excel_counter)] = self.article
        self.worksheet['I' + str(self.excel_counter)] = self.barcode
        self.worksheet['J' + str(self.excel_counter)] = self.country
        self.worksheet['K' + str(self.excel_counter)] = self.trademark
        self.worksheet['L' + str(self.excel_counter)] = self.maker
        self.worksheet['M' + str(self.excel_counter)] = self.weight
        self.worksheet['N' + str(self.excel_counter)] = self.unit

        self.worksheet['O' + str(self.excel_counter)] = self.price
        self.worksheet['P' + str(self.excel_counter)] = self.old_price
        self.worksheet['Q' + str(self.excel_counter)] = self.characteristics
        self.worksheet['R' + str(self.excel_counter)] = self.property
        self.worksheet['S' + str(self.excel_counter)] = self.about
        self.worksheet['T' + str(self.excel_counter)] = self.photo
        self.worksheet['U' + str(self.excel_counter)] = url
        self.worksheet['V' + str(self.excel_counter)] = str(datetime.datetime.now().strftime("%d-%m-%Y"))
        print(self.excel_counter)





if __name__ == '__main__':
    parser = Parser()
    with open('products.txt', 'r') as f:
        url_list = parser.try_make_new_excelbook()
        for url in f:
            url = url.replace('\n', '')
            if url in url_list:
                url_list.remove(url)
            else:
                try:
                    parser.start(url)
                except Exception as e:
                    if parser.is_remove:
                        pass
                    else:
                        print(e)
                        parser.workbook.save(path + parser.excel_filename)
                        print('saved')
                        os.startfile(
                            r'D:\python\e-dostavka.by\parser_pages.bat')
                        parser.driver.quit()
                        quit()
        parser.workbook.save(path + parser.excel_filename)
        print('saved2')





