import requests
from bs4 import BeautifulSoup
import openpyxl
import xlsxwriter
import datetime
import os

path = 'D:\python\e-dostavka.by\\'

class Parser:
    def __init__(self):
        self.excel_filename = 'products21.06.18.xlsx'
        self.excel_counter = 2
        self.workbook = openpyxl.load_workbook(path + self.excel_filename)
        self.worksheet = self.workbook['goods']


    def get_price(self):
        try:
            # print('цена self.price до: ', self.price)
            self.price = self.driver.find('div',
                                          class_='product_card__inner').find(
                'div', class_='price').text.split('\n')[0]
            # print('цена self.price после: ', self.price)
            # print('цена self.old_price до: ', self.old_price)

            self.old_price = self.driver.find(class_ = 'product_card__inner').find(class_ = 'old_price').text
            # print('цена self.old_price после: ', self.old_price)
        except AttributeError:
            self.old_price = ''

    def start(self):

        cell = self.worksheet.cell(row=self.excel_counter, column=21).value

        while cell:
            self.text = requests.get(cell).text
            self.driver = BeautifulSoup(self.text, 'lxml')
            if len(self.driver.find_all(class_='content')) == 1:
                self.is_remove = True

            self.price = 'None'
            self.old_price = ''
            self.get_price()
            cell = self.worksheet.cell(row=self.excel_counter, column=21).value
            self.worksheet['O' + str(self.excel_counter)] = self.price
            self.worksheet['P' + str(self.excel_counter)] = self.old_price
            self.worksheet['V' + str(self.excel_counter)] = str(datetime.datetime.now().strftime("%d-%m-%Y"))
            print(self.excel_counter)
            self.excel_counter += 1
        parser.workbook.save(path + parser.excel_filename)
        print('renew finished')

def worker():
    try:
        parser.start()
    except Exception as e:
        print(e)
        parser.excel_counter+=1
        worker()


if __name__ == '__main__':
    parser = Parser()
    worker()


