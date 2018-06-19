import openpyxl

path = 'D:\crawler_freelance.ru\e-dostavka.by\\'
excel_filename = 'products.xlsx'
excel_counter = 2
workbook = openpyxl.load_workbook(path + excel_filename)
worksheet = workbook['goods']
cell = worksheet.cell(row=excel_counter, column=12).value
while cell:
    worksheet['L'+ str(excel_counter)] = cell.split(' ')[0]
    worksheet['M' + str(excel_counter)] = cell.split(' ')[-1]
    excel_counter += 1
    cell = worksheet.cell(row=excel_counter, column=12).value
workbook.save(path + excel_filename)